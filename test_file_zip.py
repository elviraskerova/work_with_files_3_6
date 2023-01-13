import zipfile
from os.path import basename
from zipfile import ZipFile
import os
from PyPDF2 import PdfReader
from openpyxl.reader.excel import load_workbook

path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')
path_resources = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
my_zip_file = os.path.join(path_resources, 'info.zip')


 # создание зип
def test_create_archive():
    files = os.listdir(path)
    with ZipFile('resources/info.zip', 'w') as zip_file:
        for file in files:
            file_name = os.path.join(path, file)
            zip_file.write(file_name, basename(file_name))

#распаковка зип
# extract_dir = 'extract_dir'
#
# with zipfile.ZipFile('resources/info.zip') as zf:
#     zf.extractall(extract_dir)
#
# for file in glob.glob(extract_dir + '/**', recursive=True):
#     print(file)


    # Чтение и проверка pdf
def test_read_pdf():
    with ZipFile(my_zip_file, 'r') as zf:
        r_pdf = zf.extract('PDF_file.pdf')
        reader = PdfReader(r_pdf)
        text = reader.pages[0].extract_text()
        print(text)
        assert 'Время выполнения' in text
        os.remove(r_pdf)


    # Чтение и проверка csv
def test_read_csv():
    zip_file = ZipFile(my_zip_file)
    text = str(zip_file.read('csv_file.csv'))
    print(text)
    assert 'Da Man' in text


    # Чтение и проверка xlsx
def test_read_xlsx():
    zip_file = ZipFile(my_zip_file)
    workbook = zip_file.extract('xlsx_file.xlsx')
    xlsx_book = load_workbook(workbook)
    sheet = xlsx_book.active
    print(sheet.cell(row=1, column=1).value)
    assert 'www' in sheet.cell(row=1, column=1).value
    os.remove(workbook)





